from dataclasses import dataclass, fields
from typing import Generator, Iterable
from openpyxl import load_workbook
import itertools
import math

@dataclass
class Resistor:
    row: int
    resistance: float
    power: float
    type: str
    accuracy: float
    shipment: str

def parse_workbook(path: str) -> Generator[Resistor]:
    sheet = load_workbook(path).active
    col_mapping = [ (field, idx)
        for idx, col in enumerate(sheet[1])
        if (field := next((f for f in fields(Resistor) if col.value and f.name in col.value.lower()), None)) 
    ]
    return ( 
        Resistor(**({ field.name: field.type(row[index].value) for field, index in col_mapping } | { "row": index })) 
        for index, row in enumerate(sheet.iter_rows(min_row=2), 2) 
    )

@dataclass
class VoltageDivider:
    factor: float
    error: float
    current: float
    r_total: float
    v_tev: float
    r_tev: float
    r1: Resistor
    r2: Resistor

def find_dividers(resistors: Iterable[Resistor], v_ref: float, search_factor: float, max_error: float):
    results: list[VoltageDivider] = []
    for r1, r2 in itertools.product(resistors, repeat=2):
        r_total = r1.resistance + r2.resistance
        if r_total <= 0:
            continue
        
        factor = r2.resistance / r_total
        error = math.fabs(factor - search_factor)
        if error > max_error:
            continue

        current = v_ref / r_total
        v_tev = v_ref * factor
        r_tev = r1.resistance * r2.resistance / r_total
        results.append(VoltageDivider(factor, error, current, r_total, v_tev, r_tev, r1, r2))
        
    results.sort(key = lambda r: (r.error, r.r_tev))
    return results